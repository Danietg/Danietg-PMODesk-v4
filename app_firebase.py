import streamlit as st
import firebase_admin
from firebase_admin import credentials, db
import json
from datetime import datetime
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Configurar página
st.set_page_config(page_title="PMODesk v4.0", layout="wide", initial_sidebar_state="expanded")

# Inicializar Firebase
try:
    if not firebase_admin.get_app():
        cred_dict = st.secrets["firebase"]
        cred = credentials.Certificate(cred_dict)
        firebase_admin.initialize_app(cred, {
            'databaseURL': 'https://pmodesk-18f54-default-rtdb.firebaseio.com'
        })
except ValueError:
    pass

ref = db.reference()

# ====== FUNÇÕES AUXILIARES ======
def salvar_projeto(projeto):
    """Salvar projeto no Firebase"""
    projeto_id = str(datetime.now().timestamp()).replace('.', '')
    ref.child('projetos').child(projeto_id).set(projeto)
    return projeto_id

def listar_projetos():
    """Listar todos os projetos"""
    projetos = ref.child('projetos').get()
    if projetos.val() is None:
        return {}
    return projetos.val() if isinstance(projetos.val(), dict) else {}

def atualizar_projeto(projeto_id, projeto):
    """Atualizar projeto"""
    ref.child('projetos').child(projeto_id).set(projeto)

def deletar_projeto(projeto_id):
    """Deletar projeto"""
    ref.child('projetos').child(projeto_id).delete()

def salvar_tarefa(projeto_id, tarefa):
    """Salvar tarefa"""
    tarefa_id = str(datetime.now().timestamp()).replace('.', '')
    ref.child('tarefas').child(projeto_id).child(tarefa_id).set(tarefa)
    return tarefa_id

def listar_tarefas(projeto_id):
    """Listar tarefas de um projeto"""
    tarefas = ref.child('tarefas').child(projeto_id).get()
    if tarefas.val() is None:
        return {}
    return tarefas.val() if isinstance(tarefas.val(), dict) else {}

def deletar_tarefa(projeto_id, tarefa_id):
    """Deletar tarefa"""
    ref.child('tarefas').child(projeto_id).child(tarefa_id).delete()

def salvar_risco(projeto_id, risco):
    """Salvar risco"""
    risco_id = str(datetime.now().timestamp()).replace('.', '')
    ref.child('riscos').child(projeto_id).child(risco_id).set(risco)
    return risco_id

def listar_riscos(projeto_id):
    """Listar riscos"""
    riscos = ref.child('riscos').child(projeto_id).get()
    if riscos.val() is None:
        return {}
    return riscos.val() if isinstance(riscos.val(), dict) else {}

def deletar_risco(projeto_id, risco_id):
    """Deletar risco"""
    ref.child('riscos').child(projeto_id).child(risco_id).delete()

def exportar_excel(projetos_selecionados):
    """Exportar para Excel"""
    wb = Workbook()
    wb.remove(wb.active)
    
    for projeto_id, projeto in projetos_selecionados.items():
        ws = wb.create_sheet(projeto.get('nome', 'Projeto')[:31])
        
        # Cabeçalho
        row = 1
        headers = ['Campo', 'Valor']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row = 2
        for key, value in projeto.items():
            ws.cell(row, 1, key)
            ws.cell(row, 2, str(value))
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
    
    # Salvar em memória
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ====== SIDEBAR ======
st.sidebar.title("📊 PMODesk v4.0")
pagina = st.sidebar.radio("Menu", [
    "📊 Dashboard",
    "📁 Projetos",
    "✅ Tarefas",
    "⚠️ Riscos",
    "🔄 Mudanças",
    "📅 Comunicações",
    "💾 Backup"
])

# ====== DASHBOARD ======
if pagina == "📊 Dashboard":
    st.title("📊 Dashboard")
    
    projetos = listar_projetos()
    total_projetos = len(projetos)
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total de Projetos", total_projetos)
    with col2:
        st.metric("Projetos Ativos", len([p for p in projetos.values() if p.get('status') == 'Ativo']))
    with col3:
        st.metric("Progresso Médio", f"{sum([p.get('progresso_realizado', 0) for p in projetos.values()]) // max(total_projetos, 1)}%")
    with col4:
        st.metric("Últimas 24h", total_projetos)
    
    st.subheader("Projetos Recentes")
    if projetos:
        df = pd.DataFrame([
            {
                'Nome': p.get('nome', ''),
                'Status': p.get('status', ''),
                'Progresso': f"{p.get('progresso_realizado', 0)}%",
                'GP': p.get('gp', '')
            }
            for p in list(projetos.values())[:5]
        ])
        st.dataframe(df, use_container_width=True)

# ====== PROJETOS ======
elif pagina == "📁 Projetos":
    st.title("📁 Projetos")
    
    tab1, tab2, tab3 = st.tabs(["📋 Listar", "➕ Criar", "📊 Exportar"])
    
    # TAB: Listar
    with tab1:
        projetos = listar_projetos()
        if projetos:
            for proj_id, projeto in projetos.items():
                with st.expander(f"📌 {projeto.get('nome', 'Sem nome')}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Status:** {projeto.get('status', '')}")
                        st.write(f"**Fase:** {projeto.get('fase', '')}")
                        st.write(f"**GP:** {projeto.get('gp', '')}")
                    with col2:
                        st.write(f"**Progresso:** {projeto.get('progresso_realizado', 0)}%")
                        st.write(f"**Razão Social:** {projeto.get('razao_social', '')}")
                    
                    if st.button("🗑️ Deletar", key=f"del_{proj_id}"):
                        deletar_projeto(proj_id)
                        st.rerun()
        else:
            st.info("Nenhum projeto criado ainda!")
    
    # TAB: Criar
    with tab2:
        st.subheader("Criar Novo Projeto")
        
        with st.form("form_projeto"):
            nome = st.text_input("Nome do Projeto *")
            razao_social = st.text_input("Razão Social")
            status = st.selectbox("Status", ["Planejamento", "Execução", "Monitoramento", "Encerramento", "Ativo"])
            fase = st.selectbox("Fase", ["Iniciação", "Planejamento", "Execução", "Monitoramento", "Encerramento"])
            gp = st.text_input("Gerente de Projeto")
            gf = st.text_input("Gerente Financeiro")
            gt = st.text_input("Gerente Técnico")
            progresso = st.slider("Progresso Realizado (%)", 0, 100, 0)
            pontos_criticos = st.text_area("Pontos Críticos")
            
            if st.form_submit_button("💾 Salvar Projeto"):
                if nome:
                    projeto = {
                        'nome': nome,
                        'razao_social': razao_social,
                        'status': status,
                        'fase': fase,
                        'gp': gp,
                        'gf': gf,
                        'gt': gt,
                        'progresso_realizado': progresso,
                        'pontos_criticos': pontos_criticos,
                        'data_criacao': datetime.now().isoformat()
                    }
                    salvar_projeto(projeto)
                    st.success("✅ Projeto criado com sucesso!")
                    st.rerun()
                else:
                    st.error("Nome do projeto é obrigatório!")
    
    # TAB: Exportar
    with tab3:
        st.subheader("Exportar para Excel")
        projetos = listar_projetos()
        
        if projetos:
            selecionados = st.multiselect(
                "Selecione projetos para exportar:",
                options=list(projetos.keys()),
                format_func=lambda x: projetos[x].get('nome', x)
            )
            
            if st.button("📊 Gerar Excel"):
                if selecionados:
                    projetos_exp = {pid: projetos[pid] for pid in selecionados}
                    buffer = exportar_excel(projetos_exp)
                    
                    st.download_button(
                        label="📥 Baixar Excel",
                        data=buffer,
                        file_name=f"PMODesk_Projetos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.warning("Selecione pelo menos um projeto!")
        else:
            st.info("Nenhum projeto para exportar!")

# ====== TAREFAS ======
elif pagina == "✅ Tarefas":
    st.title("✅ Tarefas")
    
    projetos = listar_projetos()
    if not projetos:
        st.warning("Crie um projeto primeiro!")
    else:
        projeto_selecionado = st.selectbox(
            "Selecione um projeto:",
            options=list(projetos.keys()),
            format_func=lambda x: projetos[x].get('nome', x)
        )
        
        tab1, tab2 = st.tabs(["📋 Listar", "➕ Criar"])
        
        with tab1:
            tarefas = listar_tarefas(projeto_selecionado)
            if tarefas:
                for tar_id, tarefa in tarefas.items():
                    col1, col2 = st.columns([0.8, 0.2])
                    with col1:
                        st.write(f"**{tarefa.get('nome', 'Sem nome')}** ({tarefa.get('status', '')})")
                        st.write(f"Responsável: {tarefa.get('responsavel', '')}")
                    with col2:
                        if st.button("🗑️", key=f"del_tar_{tar_id}"):
                            deletar_tarefa(projeto_selecionado, tar_id)
                            st.rerun()
            else:
                st.info("Nenhuma tarefa criada!")
        
        with tab2:
            st.subheader("Criar Tarefa")
            with st.form("form_tarefa"):
                nome = st.text_input("Nome da Tarefa *")
                descricao = st.text_area("Descrição")
                responsavel = st.text_input("Responsável")
                status = st.selectbox("Status", ["Pendente", "Em Progresso", "Concluído", "Bloqueado"])
                prioridade = st.selectbox("Prioridade", ["Baixa", "Média", "Alta", "Crítica"])
                
                if st.form_submit_button("💾 Salvar Tarefa"):
                    if nome:
                        tarefa = {
                            'nome': nome,
                            'descricao': descricao,
                            'responsavel': responsavel,
                            'status': status,
                            'prioridade': prioridade,
                            'data_criacao': datetime.now().isoformat()
                        }
                        salvar_tarefa(projeto_selecionado, tarefa)
                        st.success("✅ Tarefa criada!")
                        st.rerun()
                    else:
                        st.error("Nome é obrigatório!")

# ====== RISCOS ======
elif pagina == "⚠️ Riscos":
    st.title("⚠️ Riscos")
    
    projetos = listar_projetos()
    if not projetos:
        st.warning("Crie um projeto primeiro!")
    else:
        projeto_selecionado = st.selectbox(
            "Selecione um projeto:",
            options=list(projetos.keys()),
            format_func=lambda x: projetos[x].get('nome', x),
            key="select_proj_risco"
        )
        
        tab1, tab2 = st.tabs(["📋 Listar", "➕ Criar"])
        
        with tab1:
            riscos = listar_riscos(projeto_selecionado)
            if riscos:
                for ris_id, risco in riscos.items():
                    col1, col2 = st.columns([0.8, 0.2])
                    with col1:
                        st.write(f"**{risco.get('descricao', 'Sem descrição')}**")
                        st.write(f"Probabilidade: {risco.get('probabilidade', '')} | Impacto: {risco.get('impacto', '')}")
                    with col2:
                        if st.button("🗑️", key=f"del_ris_{ris_id}"):
                            deletar_risco(projeto_selecionado, ris_id)
                            st.rerun()
            else:
                st.info("Nenhum risco criado!")
        
        with tab2:
            st.subheader("Criar Risco")
            with st.form("form_risco"):
                descricao = st.text_area("Descrição *")
                probabilidade = st.selectbox("Probabilidade", ["Baixa", "Média", "Alta", "Crítica"])
                impacto = st.selectbox("Impacto", ["Baixo", "Médio", "Alto", "Crítico"])
                tipo_resposta = st.selectbox("Tipo de Resposta", ["Mitigar", "Aceitar", "Evitar", "Transferir"])
                plano = st.text_area("Plano de Resposta")
                
                if st.form_submit_button("💾 Salvar Risco"):
                    if descricao:
                        risco = {
                            'descricao': descricao,
                            'probabilidade': probabilidade,
                            'impacto': impacto,
                            'tipo_resposta': tipo_resposta,
                            'plano': plano,
                            'data_criacao': datetime.now().isoformat()
                        }
                        salvar_risco(projeto_selecionado, risco)
                        st.success("✅ Risco criado!")
                        st.rerun()
                    else:
                        st.error("Descrição é obrigatória!")

# ====== OUTROS ======
elif pagina == "🔄 Mudanças":
    st.title("🔄 Mudanças")
    st.info("Funcionalidade em desenvolvimento...")

elif pagina == "📅 Comunicações":
    st.title("📅 Comunicações")
    st.info("Funcionalidade em desenvolvimento...")

elif pagina == "💾 Backup":
    st.title("💾 Backup")
    st.success("✅ Firebase Realtime Database faz backup automático!")
    st.info("Seus dados estão salvos na nuvem do Google Cloud!")

st.sidebar.markdown("---")
st.sidebar.write("**PMODesk v4.0**")
st.sidebar.write("Sistema de Gestão de Projetos PMBOK")
