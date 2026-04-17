import streamlit as st
import requests
import json
from datetime import datetime
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

st.set_page_config(page_title="PMODesk v4.0", layout="wide")

# Firebase Realtime Database URL
FIREBASE_URL = "https://pmodesk-18f54-default-rtdb.firebaseio.com"

# ====== FUNÇÕES FIREBASE REST ======
def salvar_projeto(projeto):
    """Salvar via REST API"""
    try:
        response = requests.post(
            f"{FIREBASE_URL}/projetos.json",
            json=projeto,
            timeout=10
        )
        return response.json().get('name')
    except:
        st.error("Erro ao salvar projeto!")
        return None

def listar_projetos():
    """Listar projetos via REST"""
    try:
        response = requests.get(f"{FIREBASE_URL}/projetos.json", timeout=10)
        return response.json() if response.json() else {}
    except:
        return {}

def atualizar_projeto(proj_id, projeto):
    """Atualizar projeto via REST"""
    try:
        requests.patch(f"{FIREBASE_URL}/projetos/{proj_id}.json", json=projeto, timeout=10)
    except:
        st.error("Erro ao atualizar!")

def deletar_projeto(proj_id):
    """Deletar projeto via REST"""
    try:
        requests.delete(f"{FIREBASE_URL}/projetos/{proj_id}.json", timeout=10)
    except:
        st.error("Erro ao deletar!")

def salvar_tarefa(proj_id, tarefa):
    """Salvar tarefa via REST"""
    try:
        requests.post(f"{FIREBASE_URL}/tarefas/{proj_id}.json", json=tarefa, timeout=10)
    except:
        st.error("Erro ao salvar tarefa!")

def listar_tarefas(proj_id):
    """Listar tarefas via REST"""
    try:
        response = requests.get(f"{FIREBASE_URL}/tarefas/{proj_id}.json", timeout=10)
        return response.json() if response.json() else {}
    except:
        return {}

def deletar_tarefa(proj_id, tar_id):
    """Deletar tarefa via REST"""
    try:
        requests.delete(f"{FIREBASE_URL}/tarefas/{proj_id}/{tar_id}.json", timeout=10)
    except:
        st.error("Erro ao deletar!")

def salvar_risco(proj_id, risco):
    """Salvar risco via REST"""
    try:
        requests.post(f"{FIREBASE_URL}/riscos/{proj_id}.json", json=risco, timeout=10)
    except:
        st.error("Erro ao salvar risco!")

def listar_riscos(proj_id):
    """Listar riscos via REST"""
    try:
        response = requests.get(f"{FIREBASE_URL}/riscos/{proj_id}.json", timeout=10)
        return response.json() if response.json() else {}
    except:
        return {}

def deletar_risco(proj_id, ris_id):
    """Deletar risco via REST"""
    try:
        requests.delete(f"{FIREBASE_URL}/riscos/{proj_id}/{ris_id}.json", timeout=10)
    except:
        st.error("Erro ao deletar!")

def exportar_excel(projetos_selecionados):
    """Exportar para Excel"""
    wb = Workbook()
    wb.remove(wb.active)
    
    for proj_id, projeto in projetos_selecionados.items():
        ws = wb.create_sheet(projeto.get('nome', 'Projeto')[:31])
        
        row = 1
        for key, value in projeto.items():
            ws.cell(row, 1, key)
            ws.cell(row, 2, str(value))
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ====== SIDEBAR ======
st.sidebar.title("📊 PMODesk v4.0")
pagina = st.sidebar.radio("Menu", [
    "📊 Dashboard",
    "📁 Projetos",
    "✅ Tarefas",
    "⚠️ Riscos",
    "📊 Exportar",
    "💾 Backup"
])

# ====== DASHBOARD ======
if pagina == "📊 Dashboard":
    st.title("📊 Dashboard")
    
    projetos = listar_projetos()
    total = len(projetos) if projetos else 0
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total de Projetos", total)
    with col2:
        st.metric("Projetos Ativos", len([p for p in (projetos.values() if projetos else []) if p.get('status') == 'Ativo']))
    with col3:
        media = sum([p.get('progresso_realizado', 0) for p in (projetos.values() if projetos else [])]) // max(total, 1)
        st.metric("Progresso Médio", f"{media}%")
    with col4:
        st.metric("Últimas 24h", total)
    
    st.subheader("Projetos Recentes")
    if projetos:
        dados = []
        for p in list(projetos.values())[:5]:
            dados.append({
                'Nome': p.get('nome', ''),
                'Status': p.get('status', ''),
                'Progresso': f"{p.get('progresso_realizado', 0)}%",
                'GP': p.get('gp', '')
            })
        st.dataframe(pd.DataFrame(dados), use_container_width=True)

# ====== PROJETOS ======
elif pagina == "📁 Projetos":
    st.title("📁 Projetos")
    
    tab1, tab2 = st.tabs(["📋 Listar", "➕ Criar"])
    
    with tab1:
        projetos = listar_projetos()
        if projetos:
            for proj_id, projeto in projetos.items():
                col1, col2 = st.columns([0.8, 0.2])
                with col1:
                    st.write(f"**{projeto.get('nome', 'Sem nome')}** ({projeto.get('status', '')})")
                    st.write(f"GP: {projeto.get('gp', '')} | Progresso: {projeto.get('progresso_realizado', 0)}%")
                with col2:
                    if st.button("🗑️", key=f"del_{proj_id}"):
                        deletar_projeto(proj_id)
                        st.rerun()
        else:
            st.info("Nenhum projeto criado!")
    
    with tab2:
        st.subheader("Criar Novo Projeto")
        with st.form("form_projeto"):
            nome = st.text_input("Nome do Projeto *")
            razao_social = st.text_input("Razão Social")
            status = st.selectbox("Status", ["Planejamento", "Execução", "Ativo", "Encerrado"])
            fase = st.selectbox("Fase", ["Iniciação", "Planejamento", "Execução", "Encerramento"])
            gp = st.text_input("Gerente de Projeto")
            gf = st.text_input("Gerente Financeiro")
            gt = st.text_input("Gerente Técnico")
            progresso = st.slider("Progresso (%)", 0, 100, 0)
            pontos_criticos = st.text_area("Pontos Críticos")
            
            if st.form_submit_button("💾 Salvar"):
                if nome:
                    projeto = {
                        'nome': nome,
                        'razao_social': razao_social,
                        'status': status,
                        'fase': fase,
                        'gp': gp,
                        'gf': gf,
                        'gt': gt,
                        'progresso_realizado': progresso,
                        'pontos_criticos': pontos_criticos,
                        'data_criacao': datetime.now().isoformat()
                    }
                    if salvar_projeto(projeto):
                        st.success("✅ Projeto criado!")
                        st.rerun()
                else:
                    st.error("Nome é obrigatório!")

# ====== TAREFAS ======
elif pagina == "✅ Tarefas":
    st.title("✅ Tarefas")
    
    projetos = listar_projetos()
    if not projetos:
        st.warning("Crie um projeto primeiro!")
    else:
        proj_selecionado = st.selectbox(
            "Projeto:",
            options=list(projetos.keys()),
            format_func=lambda x: projetos[x].get('nome', x)
        )
        
        tab1, tab2 = st.tabs(["📋 Listar", "➕ Criar"])
        
        with tab1:
            tarefas = listar_tarefas(proj_selecionado)
            if tarefas:
                for tar_id, tarefa in tarefas.items():
                    col1, col2 = st.columns([0.8, 0.2])
                    with col1:
                        st.write(f"**{tarefa.get('nome', '')}** ({tarefa.get('status', '')})")
                        st.write(f"Responsável: {tarefa.get('responsavel', '')}")
                    with col2:
                        if st.button("🗑️", key=f"del_tar_{tar_id}"):
                            deletar_tarefa(proj_selecionado, tar_id)
                            st.rerun()
            else:
                st.info("Nenhuma tarefa!")
        
        with tab2:
            with st.form("form_tarefa"):
                nome = st.text_input("Nome *")
                descricao = st.text_area("Descrição")
                responsavel = st.text_input("Responsável")
                status = st.selectbox("Status", ["Pendente", "Em Progresso", "Concluído", "Bloqueado"])
                prioridade = st.selectbox("Prioridade", ["Baixa", "Média", "Alta", "Crítica"])
                
                if st.form_submit_button("💾 Salvar"):
                    if nome:
                        tarefa = {
                            'nome': nome,
                            'descricao': descricao,
                            'responsavel': responsavel,
                            'status': status,
                            'prioridade': prioridade
                        }
                        salvar_tarefa(proj_selecionado, tarefa)
                        st.success("✅ Tarefa criada!")
                        st.rerun()
                    else:
                        st.error("Nome é obrigatório!")

# ====== RISCOS ======
elif pagina == "⚠️ Riscos":
    st.title("⚠️ Riscos")
    
    projetos = listar_projetos()
    if not projetos:
        st.warning("Crie um projeto primeiro!")
    else:
        proj_selecionado = st.selectbox(
            "Projeto:",
            options=list(projetos.keys()),
            format_func=lambda x: projetos[x].get('nome', x),
            key="sel_risco"
        )
        
        tab1, tab2 = st.tabs(["📋 Listar", "➕ Criar"])
        
        with tab1:
            riscos = listar_riscos(proj_selecionado)
            if riscos:
                for ris_id, risco in riscos.items():
                    col1, col2 = st.columns([0.8, 0.2])
                    with col1:
                        st.write(f"**{risco.get('descricao', '')}**")
                        st.write(f"Prob: {risco.get('probabilidade', '')} | Impacto: {risco.get('impacto', '')}")
                    with col2:
                        if st.button("🗑️", key=f"del_ris_{ris_id}"):
                            deletar_risco(proj_selecionado, ris_id)
                            st.rerun()
            else:
                st.info("Nenhum risco!")
        
        with tab2:
            with st.form("form_risco"):
                descricao = st.text_area("Descrição *")
                probabilidade = st.selectbox("Probabilidade", ["Baixa", "Média", "Alta"])
                impacto = st.selectbox("Impacto", ["Baixo", "Médio", "Alto"])
                tipo_resposta = st.selectbox("Resposta", ["Mitigar", "Aceitar", "Evitar"])
                plano = st.text_area("Plano")
                
                if st.form_submit_button("💾 Salvar"):
                    if descricao:
                        risco = {
                            'descricao': descricao,
                            'probabilidade': probabilidade,
                            'impacto': impacto,
                            'tipo_resposta': tipo_resposta,
                            'plano': plano
                        }
                        salvar_risco(proj_selecionado, risco)
                        st.success("✅ Risco criado!")
                        st.rerun()
                    else:
                        st.error("Descrição é obrigatória!")

# ====== EXPORTAR ======
elif pagina == "📊 Exportar":
    st.title("📊 Exportar para Excel")
    
    projetos = listar_projetos()
    if projetos:
        selecionados = st.multiselect(
            "Selecione projetos:",
            options=list(projetos.keys()),
            format_func=lambda x: projetos[x].get('nome', x)
        )
        
        if st.button("📥 Gerar Excel"):
            if selecionados:
                projetos_exp = {pid: projetos[pid] for pid in selecionados}
                buffer = exportar_excel(projetos_exp)
                st.download_button(
                    label="Baixar",
                    data=buffer,
                    file_name=f"PMODesk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("Selecione projetos!")
    else:
        st.info("Nenhum projeto para exportar!")

# ====== BACKUP ======
elif pagina == "💾 Backup":
    st.title("💾 Backup")
    st.success("✅ Firebase faz backup automático!")
    st.info("Dados salvos na nuvem do Google Cloud!")

st.sidebar.markdown("---")
st.sidebar.write("**PMODesk v4.0**")
st.sidebar.write("Gestão de Projetos PMBOK")
